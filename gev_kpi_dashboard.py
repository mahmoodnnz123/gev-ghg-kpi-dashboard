"""
GE Vernova — GHG Emissions Trend & Sustainability KPI Dashboard
Project 1 of 2

Matin Bahadori, 2025
M.Sc. Umwelttechnik — BTU Cottbus-Senftenberg

Data source : GE Vernova Sustainability Report 2025, p. 97
Standard    : GHG Protocol (market-based Scope 2)
Target      : Carbon neutrality by 2030 (Scope 1+2)
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib.patches import FancyBboxPatch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

DARK_TEXT = "#1A1A1A"
MID_TEXT  = "#666666"
GRID_COL  = "#E8E8E8"
GEV_BLUE  = "#005EB8"
GEV_GREEN = "#00A651"
GEV_RED   = "#E53935"
GEV_ORANGE= "#FF6F00"

plt.rcParams.update({
    "font.family":       "DejaVu Sans",
    "font.size":         10,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.edgecolor":    "#444444",
    "axes.linewidth":    0.8,
})


# ════════════════════════════════════════════════════════════════════════════
# DATA — GE Vernova Sustainability Report 2025, p. 97
# ════════════════════════════════════════════════════════════════════════════

GHG = pd.DataFrame({
    "year":         [2019,    2020,    2021,    2022,    2023,    2024,    2025],
    "scope1_s2_mb": [880_348, None,    None,    None,    544_516, 428_213, 313_659],
    "scope2_mb":    [512_753, None,    None,    None,    297_705, 201_402,  97_953],
    "scope2_lb":    [558_830, None,    None,    None,    376_537, 360_377, 314_081],
    "scope1_s2_lb": [926_425, None,    None,    None,    623_349, 587_188, 529_787],
    "pct_reduction":[0,       None,    None,    None,    38,      51,      64],
})
for col in ["scope1_s2_mb", "scope2_mb"]:
    GHG[col] = pd.to_numeric(GHG[col], errors="coerce")
    GHG[col] = GHG[col].interpolate(method="linear")

CAPACITY = pd.DataFrame({
    "year":          [2023, 2024, 2025],
    "new_gen_gw":    [29,   31,   26],
    "grid_gw":       [64,   71,   68],
    "avoided_mmt":   [15,   27,   22],
    "carbon_int":    [334,  368,  309],
    "dev_econ_pct":  [42,   62,   47],
})

print("Data loaded.")
print(f"2019 baseline : 880,348 tCO2e")
print(f"2025 actual   : 313,659 tCO2e  (−64%)")
print(f"2030 target   : 0 tCO2e (carbon neutral)")


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 1 — GHG TRAJECTORY + NET ZERO MODELLING
# ════════════════════════════════════════════════════════════════════════════

fig1 = plt.figure(figsize=(18, 12))
fig1.patch.set_facecolor("white")
fig1.suptitle(
    "Figure 1.  GHG Emissions Trend & Net Zero Trajectory — GE Vernova (2019–2030)",
    fontsize=14, fontweight="bold", y=0.98, color=DARK_TEXT)
fig1.text(0.5, 0.945,
    "Data: GE Vernova Sustainability Report 2025, p. 97  |  "
    "GHG Protocol market-based Scope 2  |  "
    "Baseline: 880,348 tCO₂e (2019)  |  Target: Carbon neutrality 2030",
    ha="center", fontsize=9, color=MID_TEXT, style="italic")

gs = gridspec.GridSpec(2, 3, figure=fig1,
                       top=0.91, bottom=0.07,
                       hspace=0.45, wspace=0.35)

# ── Panel A: Absolute emissions + trajectories ────────────────────────────────
ax1 = fig1.add_subplot(gs[0, :2])
ax1.set_facecolor("#F8F8F8")

# actual reported points
reported_yrs  = [2019, 2023, 2024, 2025]
reported_vals = [880_348, 544_516, 428_213, 313_659]
ax1.plot(GHG["year"], GHG["scope1_s2_mb"] / 1000,
         color=GEV_BLUE, linewidth=2.2, marker="o",
         markersize=6, label="Actual Scope 1+2 (MB)", zorder=5)

# dashed interpolated segment
for i in [1, 2, 3]:
    ax1.plot([GHG["year"].iloc[i-1], GHG["year"].iloc[i]],
             [GHG["scope1_s2_mb"].iloc[i-1] / 1000,
              GHG["scope1_s2_mb"].iloc[i] / 1000],
             color=GEV_BLUE, linewidth=1.5,
             linestyle="--", alpha=0.4, zorder=4)
ax1.text(2021, GHG.loc[GHG.year == 2021, "scope1_s2_mb"].values[0] / 1000 + 25,
         "estimated (interpolated)",
         fontsize=7.5, color=GEV_BLUE, alpha=0.55, ha="center")

# future trajectories
future_yrs = np.array([2025, 2026, 2027, 2028, 2029, 2030])
v_2025     = 313_659 / 1000

linear_path = np.linspace(v_2025, 0, len(future_yrs))
ax1.plot(future_yrs, linear_path, color=GEV_GREEN,
         linewidth=2, linestyle="--",
         label="Linear to net zero (2030)", zorder=4)

pct_per_yr = (428_213 - 313_659) / 428_213
exp_path   = [v_2025 * (1 - pct_per_yr)**i for i in range(len(future_yrs))]
ax1.plot(future_yrs, exp_path, color=GEV_ORANGE,
         linewidth=2, linestyle=":",
         label="At current rate (~27%/yr)", zorder=4)

sbti_path = [880_348 / 1000 * (1 - 0.042)**i for i in range(12)]
ax1.plot(range(2019, 2031), sbti_path, color="#9C27B0",
         linewidth=1.5, linestyle="-.", alpha=0.7,
         label="SBTi 1.5°C path (−4.2%/yr)", zorder=3)

ax1.scatter([2030], [0], color=GEV_GREEN, s=200, zorder=6,
            marker="*", label="Carbon neutrality target (2030)")
ax1.annotate("Target: 0\n(carbon neutral 2030)",
             xy=(2030, 0), xytext=(2027.5, 100),
             fontsize=9, color=GEV_GREEN, fontweight="bold",
             arrowprops=dict(arrowstyle="->", color=GEV_GREEN))

for yr, val, lbl in [(2019, 880, "Baseline\n880k t"),
                     (2025, 314, "2025\n314k t\n(−64%)")]:
    ax1.annotate(lbl, xy=(yr, val), xytext=(yr, val + 70),
                 fontsize=8, color=DARK_TEXT, ha="center",
                 arrowprops=dict(arrowstyle="-", color="#aaaaaa", lw=0.8))

ax1.set_ylabel("Scope 1+2 Emissions (thousand tCO₂e)", fontsize=10)
ax1.set_xlabel("Year", fontsize=10)
ax1.set_ylim(-50, 1050)
ax1.set_xticks(range(2019, 2031))
ax1.set_xticklabels(range(2019, 2031), rotation=30)
ax1.yaxis.grid(True, color=GRID_COL, linewidth=0.7, zorder=0)
ax1.set_axisbelow(True)
ax1.legend(frameon=False, fontsize=9, loc="upper right")
ax1.set_title("(A)  Scope 1+2 GHG Emissions — Actual vs. Net Zero Trajectories",
              fontsize=11, fontweight="bold", pad=10, loc="left", color=DARK_TEXT)

# ── Panel B: Progress gauge ────────────────────────────────────────────────────
ax2 = fig1.add_subplot(gs[0, 2])
ax2.set_facecolor("white")
ax2.axis("off")
ax2.set_title("(B)  Reduction Progress vs. 2030 Target",
              fontsize=11, fontweight="bold", pad=10, loc="left", color=DARK_TEXT)

milestones = [(38, "2023  −38%", "#FFE082"),
              (51, "2024  −51%", "#FFB300"),
              (64, "2025  −64%", "#00A651"),
              (100,"2030  Target", "#1B5E20")]
for i, (pct, lbl, c) in enumerate(milestones):
    y = 0.78 - i * 0.18
    bg = FancyBboxPatch((0.05, y), 0.85, 0.12,
                        boxstyle="round,pad=0.01",
                        facecolor="#EEEEEE", edgecolor="white",
                        transform=ax2.transAxes)
    bar = FancyBboxPatch((0.05, y), pct / 100 * 0.85, 0.12,
                         boxstyle="round,pad=0.01",
                         facecolor=c, edgecolor="white",
                         transform=ax2.transAxes, alpha=0.85)
    ax2.add_patch(bg); ax2.add_patch(bar)
    ax2.text(0.02, y + 0.06, lbl, va="center", fontsize=9,
             fontweight="bold", color=DARK_TEXT, transform=ax2.transAxes)
    ax2.text(0.93, y + 0.06, f"{pct}%", va="center", fontsize=10,
             fontweight="bold", color=c if c != "#FFE082" else "#555",
             transform=ax2.transAxes)

ax2.text(0.5, 0.08,
         "36% still needed\nby 2030 in 5 years",
         ha="center", fontsize=9, color=GEV_RED,
         fontweight="bold", transform=ax2.transAxes)

# ── Panel C: Scope 2 MB vs LB ─────────────────────────────────────────────────
ax3 = fig1.add_subplot(gs[1, :2])
ax3.set_facecolor("#F8F8F8")

yrs4   = [2019, 2023, 2024, 2025]
s2_mb  = [512_753, 297_705, 201_402,  97_953]
s2_lb  = [558_830, 376_537, 360_377, 314_081]

ax3.plot(yrs4, [v / 1000 for v in s2_mb], color=GEV_BLUE,
         linewidth=2.2, marker="o", markersize=7,
         label="Market-based (renewable certificates)", zorder=5)
ax3.plot(yrs4, [v / 1000 for v in s2_lb], color=GEV_RED,
         linewidth=2.2, marker="s", markersize=7, linestyle="--",
         label="Location-based (grid average)", zorder=5)
ax3.fill_between(yrs4, [v / 1000 for v in s2_mb],
                 [v / 1000 for v in s2_lb],
                 alpha=0.12, color=GEV_BLUE,
                 label="Gap = value of green energy procurement")

mb_drop = (512_753 - 97_953) / 512_753 * 100
ax3.annotate(f"Market-based: −{mb_drop:.0f}%\n(green energy purchases)",
             xy=(2025, 97_953 / 1000), xytext=(2023, 60),
             fontsize=8.5, color=GEV_BLUE, fontweight="bold",
             arrowprops=dict(arrowstyle="->", color=GEV_BLUE))
ax3.set_ylabel("Scope 2 Emissions (thousand tCO₂e)", fontsize=10)
ax3.set_xlabel("Year", fontsize=10)
ax3.set_ylim(0, 650)
ax3.set_xticks(yrs4)
ax3.yaxis.grid(True, color=GRID_COL, linewidth=0.7, zorder=0)
ax3.set_axisbelow(True)
ax3.legend(frameon=False, fontsize=9)
ax3.set_title("(C)  Scope 2 — Market-Based vs. Location-Based\n"
              "Gap shows value of renewable energy procurement (PPAs, RECs)",
              fontsize=11, fontweight="bold", pad=10, loc="left", color=DARK_TEXT)

# ── Panel D: Avoided vs operational ───────────────────────────────────────────
ax4 = fig1.add_subplot(gs[1, 2])
ax4.set_facecolor("#F8F8F8")

x = np.arange(3)
w = 0.35
op_kt  = [544, 428, 314]
av_mmt = [15,  27,  22]

b1 = ax4.bar(x - w/2, op_kt,              w, color=GEV_BLUE,
             edgecolor="white", alpha=0.85, label="Operational (kt)")
b2 = ax4.bar(x + w/2, [v * 1000 for v in av_mmt], w, color=GEV_GREEN,
             edgecolor="white", alpha=0.85, label="Avoided by products (kt)")

for bar, val in zip(b1, op_kt):
    ax4.text(bar.get_x() + bar.get_width()/2, val + 60,
             f"{val}kt", ha="center", fontsize=8,
             color=GEV_BLUE, fontweight="bold")
for bar, val in zip(b2, [v * 1000 for v in av_mmt]):
    ax4.text(bar.get_x() + bar.get_width()/2, val + 60,
             f"{val//1000}Mt", ha="center", fontsize=8,
             color=GEV_GREEN, fontweight="bold")

ratios = [av / op * 1000 for av, op in zip(av_mmt, op_kt)]
for i, r in enumerate(ratios):
    ax4.text(i, max(op_kt[i], av_mmt[i] * 1000) + 1200,
             f"Ratio:\n{r:.0f}x", ha="center", fontsize=8.5,
             color=DARK_TEXT, fontweight="bold")

ax4.set_xticks(x); ax4.set_xticklabels(["2023","2024","2025"])
ax4.set_ylabel("Emissions (thousand tCO₂e)", fontsize=9)
ax4.yaxis.grid(True, color=GRID_COL, linewidth=0.7, zorder=0)
ax4.set_axisbelow(True)
ax4.legend(frameon=False, fontsize=8)
ax4.set_title("(D)  Operational Emissions vs. CO₂ Avoided\nby GE Vernova Products",
              fontsize=11, fontweight="bold", pad=10, loc="left", color=DARK_TEXT)

fig1.text(0.01, 0.01,
    "Note: 2020–2022 interpolated (not reported by GE Vernova). "
    "Linear trajectory = equal annual cuts 2025→2030. "
    "Current rate = 27% YoY reduction observed 2024→2025. "
    "SBTi 1.5°C = 4.2% linear reduction from 2019.",
    fontsize=7.5, color=MID_TEXT, style="italic")

plt.savefig(f"{OUTPUT_DIR}/fig1_ghg_trajectory.png",
            dpi=150, bbox_inches="tight", facecolor="white")
plt.show(); plt.close()
print("Figure 1 saved.")


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 2 — FULL KPI DASHBOARD (4 PILLARS)
# ════════════════════════════════════════════════════════════════════════════

fig2, axes2 = plt.subplots(2, 3, figsize=(18, 11))
fig2.patch.set_facecolor("white")
fig2.suptitle(
    "Figure 2.  GE Vernova Sustainability KPI Dashboard — All Four Pillars (2023–2025)",
    fontsize=14, fontweight="bold", y=0.99, color=DARK_TEXT)
fig2.text(0.5, 0.955,
    "Data: GE Vernova Sustainability Report 2025, p. 97  |  "
    "Electrify · Decarbonize · Conserve · Thrive",
    ha="center", fontsize=9, color=MID_TEXT, style="italic")

yrs3  = ["2023","2024","2025"]
x_pos = np.arange(3)
W     = 0.35

def kpi_bar(ax, vals_a, vals_b, lbl_a, lbl_b, title, ylabel, ca, cb):
    ax.set_facecolor("#F8F8F8")
    b1 = ax.bar(x_pos-W/2, vals_a, W, color=ca, edgecolor="white", alpha=0.85, label=lbl_a)
    b2 = ax.bar(x_pos+W/2, vals_b, W, color=cb, edgecolor="white", alpha=0.85, label=lbl_b)
    for bar, v in zip(b1, vals_a):
        ax.text(bar.get_x()+bar.get_width()/2, v*1.03,
                f"{v:.2f}" if v < 10 else f"{v:.0f}",
                ha="center", fontsize=8, fontweight="bold", color=ca)
    for bar, v in zip(b2, vals_b):
        ax.text(bar.get_x()+bar.get_width()/2, v*1.03,
                f"{v:.2f}" if v < 10 else f"{v:.0f}",
                ha="center", fontsize=8, fontweight="bold", color=cb)
    ax.set_xticks(x_pos); ax.set_xticklabels(yrs3)
    ax.set_ylabel(ylabel, fontsize=9)
    ax.set_title(title, fontsize=10, fontweight="bold", pad=8, loc="left", color=DARK_TEXT)
    ax.yaxis.grid(True, color=GRID_COL, linewidth=0.7, zorder=0)
    ax.set_axisbelow(True)
    ax.legend(frameon=False, fontsize=8)

kpi_bar(axes2[0,0], [29,31,26], [64,71,68],
        "New Generating (GW)", "Grid Enabling (GW)",
        "ELECTRIFY — Capacity Installed (GW)", "GW",
        GEV_BLUE, GEV_GREEN)

kpi_bar(axes2[0,1], [334,368,309], [446,446,446],
        "GE Vernova (new)", "Global grid avg (IEA)",
        "DECARBONIZE — Carbon Intensity (g CO₂/kWh)", "g CO₂/kWh",
        GEV_BLUE, GEV_RED)

axes2[0,2].set_facecolor("#F8F8F8")
bars_av = axes2[0,2].bar(x_pos, [15,27,22], color=GEV_GREEN,
                          edgecolor="white", alpha=0.85, width=0.55)
for bar, v in zip(bars_av, [15,27,22]):
    axes2[0,2].text(bar.get_x()+bar.get_width()/2, v+0.3,
                    f"{v} MMT", ha="center", fontsize=9,
                    fontweight="bold", color=GEV_GREEN)
axes2[0,2].set_xticks(x_pos); axes2[0,2].set_xticklabels(yrs3)
axes2[0,2].set_ylabel("MMT CO₂ Avoided/yr", fontsize=9)
axes2[0,2].yaxis.grid(True, color=GRID_COL, linewidth=0.7, zorder=0)
axes2[0,2].set_axisbelow(True)
axes2[0,2].set_title("DECARBONIZE — CO₂ Avoided by Products",
                     fontsize=10, fontweight="bold", pad=8, loc="left", color=DARK_TEXT)

kpi_bar(axes2[1,0], [23,38,53], [36,53,76],
        "4R Circularity (%)", "LCA/EPD Coverage (%)",
        "CONSERVE — Product Circularity (%)", "%",
        GEV_ORANGE, GEV_GREEN)
axes2[1,0].axhline(100, color="#888", linewidth=0.8,
                   linestyle="--", alpha=0.5)
axes2[1,0].set_ylim(0, 115)

kpi_bar(axes2[1,1], [0.44,0.43,0.43], [0.21,0.21,0.22],
        "TRIR", "Days Away Rate",
        "THRIVE — Safety Rates (per 200k hrs)", "Rate",
        GEV_RED, GEV_ORANGE)
axes2[1,1].set_ylim(0, 0.6)

ax_t = axes2[1,2]
ax_t.set_facecolor("#F8F8F8")
ax_twin = ax_t.twinx()
b_eng = ax_t.bar(x_pos-W/2, [73,76,79], W, color=GEV_BLUE,
                  edgecolor="white", alpha=0.85, label="Engagement Score")
b_fat = ax_twin.bar(x_pos+W/2, [3,3,4], W, color=GEV_RED,
                     edgecolor="white", alpha=0.75, label="Total Fatalities")
for bar, v in zip(b_eng, [73,76,79]):
    ax_t.text(bar.get_x()+bar.get_width()/2, v+0.3,
              str(v), ha="center", fontsize=9, fontweight="bold", color=GEV_BLUE)
for bar, v in zip(b_fat, [3,3,4]):
    ax_twin.text(bar.get_x()+bar.get_width()/2, v+0.05,
                 str(v), ha="center", fontsize=9,
                 fontweight="bold", color=GEV_RED)
ax_t.set_ylabel("Engagement Score", fontsize=9, color=GEV_BLUE)
ax_twin.set_ylabel("Total Fatalities", fontsize=9, color=GEV_RED)
ax_t.set_xticks(x_pos); ax_t.set_xticklabels(yrs3)
ax_t.set_ylim(60, 90); ax_twin.set_ylim(0, 6)
ax_t.yaxis.grid(True, color=GRID_COL, linewidth=0.7, zorder=0)
ax_t.set_axisbelow(True)
ax_t.legend(handles=[mpatches.Patch(color=GEV_BLUE, label="Engagement Score"),
                     mpatches.Patch(color=GEV_RED,  label="Total Fatalities")],
            frameon=False, fontsize=8, loc="upper left")
ax_t.set_title("THRIVE — Engagement vs. Fatalities",
               fontsize=10, fontweight="bold", pad=8, loc="left", color=DARK_TEXT)

fig2.text(0.01, 0.01,
    "TRIR = Total Recordable Injury Rate per 200,000 hrs. "
    "4R = Rethink, Reduce, Reuse, Recycle. "
    "Carbon-free electricity 2025: 79%. "
    "Source: GE Vernova Sustainability Report 2025, p. 97.",
    fontsize=7.5, color=MID_TEXT, style="italic")

plt.tight_layout(rect=[0, 0.04, 1, 0.95])
plt.savefig(f"{OUTPUT_DIR}/fig2_kpi_dashboard.png",
            dpi=150, bbox_inches="tight", facecolor="white")
plt.show(); plt.close()
print("Figure 2 saved.")


# ════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ════════════════════════════════════════════════════════════════════════════

wb  = Workbook()
hdr_fill = PatternFill("solid", start_color="005EB8")
grn_fill = PatternFill("solid", start_color="E8F5E9")
red_fill = PatternFill("solid", start_color="FFEBEE")
ora_fill = PatternFill("solid", start_color="FFF3E0")
bold_wht = Font(name="Arial", bold=True, color="FFFFFF", size=10)
norm_blk = Font(name="Arial", color="1A1A1A", size=10)
center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin     = Side(style="thin", color="CCCCCC")
bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

def make_hdr(ws, headers, widths):
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(1,c,h)
        cell.font=bold_wht; cell.fill=hdr_fill
        cell.alignment=center; cell.border=bdr
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[1].height=30

# Sheet 1: GHG emissions
ws1 = wb.active; ws1.title = "GHG Emissions"
ws1.sheet_view.showGridLines = False
make_hdr(ws1, ["Year","Scope 1+2 MB (tCO2e)","Scope 2 MB (tCO2e)",
               "Scope 2 LB (tCO2e)","Scope 1+2 LB (tCO2e)",
               "% Reduction vs 2019","YoY Change"],
         [10,22,20,20,20,20,18])
ghg_data=[
    (2019,880_348,512_753,558_830,926_425,"0% (baseline)","—"),
    (2023,544_516,297_705,376_537,623_349,"−38%","—"),
    (2024,428_213,201_402,360_377,587_188,"−51%","−114,303"),
    (2025,313_659, 97_953,314_081,529_787,"−64%","−114,554"),
]
for r,(yr,s12mb,s2mb,s2lb,s12lb,pct,yoy) in enumerate(ghg_data,2):
    fill = red_fill if r==2 else (ora_fill if r==3 else grn_fill)
    for c,v in enumerate([yr,s12mb,s2mb,s2lb,s12lb,pct,yoy],1):
        cell=ws1.cell(r,c,v)
        cell.font=norm_blk; cell.fill=fill
        cell.alignment=center; cell.border=bdr

# Sheet 2: Net zero trajectory
ws2 = wb.create_sheet("Net Zero Trajectory")
ws2.sheet_view.showGridLines = False
make_hdr(ws2, ["Year","Linear Path (tCO2e)","At Current Rate (tCO2e)",
               "SBTi 1.5°C Path (tCO2e)","Required Annual Cut"],
         [10,22,22,22,22])
v = 313_659
pct_yr = 0.27
sbti_2025 = 880_348 * (1-0.042)**6
for r, yr in enumerate(range(2025,2031),2):
    i = yr - 2025
    lin = max(0, int(313_659 * (1 - i/5)))
    exp = int(313_659 * (1-pct_yr)**i)
    sbti = int(880_348 * (1-0.042)**(yr-2019))
    req  = int(313_659 / 5)
    fill = grn_fill if yr == 2030 else PatternFill()
    for c,val in enumerate([yr,lin,exp,sbti,req],1):
        cell=ws2.cell(r,c,val)
        cell.font=norm_blk; cell.fill=fill
        cell.alignment=center; cell.border=bdr

# Sheet 3: KPI tracker
ws3 = wb.create_sheet("KPI Tracker")
ws3.sheet_view.showGridLines = False
make_hdr(ws3, ["Metric","Pillar","2023","2024","2025","Target","Direction"],
         [35,14,12,12,12,16,12])
kpis=[
    ("New Generating Capacity (GW)","Electrify",29,31,26,"150 GW by 2030","↑"),
    ("Grid Enabling Capacity (GW)","Electrify",64,71,68,"—","↑"),
    ("CO2 Avoided (MMT/yr)","Decarbonize",15,27,22,"—","↑"),
    ("Carbon Intensity of new capacity (g CO2/kWh)","Decarbonize",334,368,309,"< 446","↓"),
    ("Scope 1+2 MB (tCO2e)","Decarbonize",544_516,428_213,313_659,"0 by 2030","↓"),
    ("Scope 1+2 Reduction vs 2019 (%)","Decarbonize",38,51,64,"100%","↑"),
    ("Carbon-Free Electricity (%)","Decarbonize","—","—",79,"100%","↑"),
    ("4R Circularity Coverage (%)","Conserve",23,38,53,"100%","↑"),
    ("LCA/EPD Coverage (%)","Conserve",36,53,76,"100%","↑"),
    ("TRIR","Thrive",0.44,0.43,0.43,"Industry avg","↓"),
    ("Days Away Rate","Thrive",0.21,0.21,0.22,"Industry avg","↓"),
    ("Employee Fatalities","Thrive",0,1,2,"0","↓"),
    ("Contractor Fatalities","Thrive",3,2,2,"0","↓"),
    ("Employee Engagement Score","Thrive",73,76,79,"—","↑"),
    ("Survey Participation (%)","Thrive",65,73,75,"—","↑"),
]
for r,(name,pillar,v23,v24,v25,tgt,direction) in enumerate(kpis,2):
    for c,v in enumerate([name,pillar,v23,v24,v25,tgt,direction],1):
        cell=ws3.cell(r,c,v)
        cell.font=norm_blk
        cell.alignment=center if c!=1 else left
        cell.border=bdr

path_xl = f"{OUTPUT_DIR}/gev_kpi_dashboard.xlsx"
wb.save(path_xl)
print(f"Excel saved → {path_xl}")

print(f"\n{'='*60}")
print(f"  GE Vernova — GHG & KPI Summary")
print(f"{'='*60}")
print(f"  2019 baseline          : 880,348 tCO2e")
print(f"  2025 actual            : 313,659 tCO2e (−64%)")
print(f"  Remaining to 2030 zero : 313,659 tCO2e in 5 years")
print(f"  At current 27%/yr rate : {int(313_659*(1-0.27)**5):,} tCO2e by 2030")
print(f"  Products avoid         : 22 MMT CO2/yr")
print(f"  Avoidance ratio        : {22e6/313_659:.0f}x operational footprint")
print(f"{'='*60}")

